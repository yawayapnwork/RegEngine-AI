"""Excel workbook export for the audit binder -- one sheet per data
category (rule changes, HITL approvals, ledger proof entries, source
circulars), which is what a compliance team actually does with this data
in practice: filter/pivot in Excel rather than read a PDF for row-level
work. The PDF (app.analytics.pdf_report) stays the executive-narrative
artifact; this is the working-data artifact.
"""
from __future__ import annotations

import io

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from app.analytics.models import AuditTrailEntry
from app.reporting.data_collector import HITLApprovalRecord, RuleChangeRecord, SourceCircularRecord

_HEADER_FILL = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
_HEADER_FONT = Font(color="FFFFFF", bold=True)


def _write_sheet(ws: Worksheet, headers: list[str], rows: list[tuple]) -> None:
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
    for row in rows:
        ws.append(row)

    for col_idx, header in enumerate(headers, start=1):
        column_letter = get_column_letter(col_idx)
        max_len = max([len(header)] + [len(str(row[col_idx - 1])) for row in rows]) if rows else len(header)
        ws.column_dimensions[column_letter].width = min(60, max_len + 2)
    ws.freeze_panes = "A2"


def build_excel_workbook(
    *,
    period_label: str,
    rule_changes: list[RuleChangeRecord],
    hitl_approvals: list[HITLApprovalRecord],
    ledger_entries: list[AuditTrailEntry],
    source_circulars: list[SourceCircularRecord],
) -> bytes:
    wb = Workbook()

    summary_ws = wb.active
    summary_ws.title = "Summary"
    summary_ws.append(["RegEngine AI -- SEBI Compliance Audit Binder"])
    summary_ws["A1"].font = Font(bold=True, size=14)
    summary_ws.append(["Reporting Period", period_label])
    summary_ws.append(["Rule Changes", len(rule_changes)])
    summary_ws.append(["HITL Approval/Rejection Cases", len(hitl_approvals)])
    summary_ws.append(["Ledger Entries (Transaction Proofs)", len(ledger_entries)])
    summary_ws.append(["Source Circulars Referenced", len(source_circulars)])
    for row in summary_ws.iter_rows(min_row=2, max_row=6, min_col=1, max_col=1):
        row[0].font = Font(bold=True)
    summary_ws.column_dimensions["A"].width = 36
    summary_ws.column_dimensions["B"].width = 40

    rule_ws = wb.create_sheet("Rule Changes")
    _write_sheet(
        rule_ws,
        ["Rule ID", "Version", "Clause", "Circular", "Active", "HITL Status", "Compiler Version", "OPA Package", "Created At (UTC)"],
        [
            (r.rule_id, r.rule_version, r.clause_number or "", r.circular_number, r.is_active, r.hitl_status, r.compiler_version or "", r.opa_package_name or "", r.created_at.isoformat())
            for r in rule_changes
        ],
    )

    hitl_ws = wb.create_sheet("HITL Approvals")
    _write_sheet(
        hitl_ws,
        ["Review ID", "Clause", "Circular", "Reason Code", "Severity", "Status", "Officer", "Flagged At (UTC)", "Resolved At (UTC)", "Resolution Notes"],
        [
            (
                h.review_id, h.clause_number or "", h.circular_number, h.reason_code, h.severity, h.status,
                h.compliance_officer_id or "", h.flagged_at.isoformat(), h.resolved_at.isoformat() if h.resolved_at else "",
                h.resolution_notes or "",
            )
            for h in hitl_approvals
        ],
    )

    ledger_ws = wb.create_sheet("Ledger Proof Chain")
    _write_sheet(
        ledger_ws,
        ["Sequence #", "Broker ID", "Transaction ID", "Evaluated At (UTC)", "Circular", "Section", "Rule ID", "Result", "Payload Digest (SHA-256)", "Current Hash (SHA-256)"],
        [
            (e.sequence_num, e.broker_id, e.transaction_id, e.evaluated_at.isoformat(), e.circular_id, e.section_reference, e.rule_id, e.evaluation_result, e.payload_digest, e.current_hash)
            for e in ledger_entries
        ],
    )

    circulars_ws = wb.create_sheet("Source Circulars")
    _write_sheet(
        circulars_ws,
        ["Circular Number", "Title", "Issue Date", "Source URL", "Raw Text Digest (SHA-256)", "Department"],
        [
            (c.circular_number, c.title or "", c.issue_date.isoformat() if c.issue_date else "", c.source_url or "", c.raw_text_digest, c.department or "")
            for c in source_circulars
        ],
    )

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
