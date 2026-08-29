"""Tests for the regengine-report CLI's supporting modules: period
resolution, RSA-PSS digital signatures, and the Excel/JSON export
builders. `app.reporting.audit_binder`'s DB/ledger-dependent orchestration
is not covered here (needs a live Postgres) -- these test everything that
IS pure/unit-testable in isolation, matching the depth of testing already
applied to this codebase's other pure modules (e.g. app.ledger.hash_chain)."""
from __future__ import annotations

import datetime as dt
import hashlib
import io
import json

import pytest

from app.analytics.models import AuditTrailEntry
from app.config import Settings
from app.reporting.data_collector import HITLApprovalRecord, RuleChangeRecord, SourceCircularRecord
from app.reporting.excel_export import build_excel_workbook
from app.reporting.period import resolve_fiscal_year, resolve_period, resolve_quarter
from app.reporting.signing import (
    SigningKeyNotConfiguredError,
    generate_signing_keypair,
    sign_manifest,
    verify_signature,
)


class TestPeriodResolution:
    def test_q1_is_apr_jun(self) -> None:
        p = resolve_quarter("Q1-2025")
        assert (p.start_date, p.end_date) == (dt.date(2025, 4, 1), dt.date(2025, 6, 30))

    def test_q4_rolls_into_next_calendar_year(self) -> None:
        p = resolve_quarter("Q4-2025")
        assert (p.start_date, p.end_date) == (dt.date(2026, 1, 1), dt.date(2026, 3, 31))

    def test_fiscal_year_spans_apr_to_mar(self) -> None:
        p = resolve_fiscal_year("FY2025-26")
        assert (p.start_date, p.end_date) == (dt.date(2025, 4, 1), dt.date(2026, 3, 31))

    def test_non_consecutive_fy_years_rejected(self) -> None:
        with pytest.raises(ValueError):
            resolve_fiscal_year("FY2025-28")

    def test_malformed_quarter_spec_rejected(self) -> None:
        with pytest.raises(ValueError):
            resolve_quarter("Q5-2025")
        with pytest.raises(ValueError):
            resolve_quarter("not-a-quarter")

    def test_resolve_period_requires_exactly_one_input_shape(self) -> None:
        with pytest.raises(ValueError):
            resolve_period(quarter=None, fiscal_year=None, start=None, end=None)
        with pytest.raises(ValueError):
            resolve_period(quarter="Q1-2025", fiscal_year="FY2025-26", start=None, end=None)

    def test_resolve_period_start_end_pair(self) -> None:
        p = resolve_period(quarter=None, fiscal_year=None, start="2025-07-01", end="2025-07-31")
        assert (p.start_date, p.end_date) == (dt.date(2025, 7, 1), dt.date(2025, 7, 31))

    def test_start_without_end_rejected(self) -> None:
        with pytest.raises(ValueError):
            resolve_period(quarter=None, fiscal_year=None, start="2025-07-01", end=None)


class TestDigitalSignature:
    def _settings_with_key(self) -> Settings:
        private_pem, _public_pem = generate_signing_keypair()
        return Settings(audit_binder_signing_private_key_pem=private_pem)

    def test_sign_and_verify_roundtrip(self) -> None:
        settings = self._settings_with_key()
        manifest = b'{"files": [{"path": "a.rego", "sha256": "abc"}]}'
        signature = sign_manifest(manifest, settings)
        assert signature.algorithm == "RSA-PSS-SHA256"
        assert verify_signature(manifest, signature) is True

    def test_tampered_manifest_fails_verification(self) -> None:
        settings = self._settings_with_key()
        manifest = b'{"files": [{"path": "a.rego", "sha256": "abc"}]}'
        signature = sign_manifest(manifest, settings)
        tampered = b'{"files": [{"path": "a.rego", "sha256": "TAMPERED"}]}'
        assert verify_signature(tampered, signature) is False

    def test_wrong_public_key_fails_verification(self) -> None:
        settings_a = self._settings_with_key()
        _priv_b, pub_b = generate_signing_keypair()
        manifest = b'{"files": []}'
        signature = sign_manifest(manifest, settings_a)
        signature.public_key_pem = pub_b  # substitute an unrelated key pair's public key
        assert verify_signature(manifest, signature) is False

    def test_missing_signing_key_raises(self) -> None:
        settings = Settings(audit_binder_signing_private_key_pem=None)
        with pytest.raises(SigningKeyNotConfiguredError):
            sign_manifest(b"{}", settings)

    def test_manifest_sha256_matches_recomputed_digest(self) -> None:
        settings = self._settings_with_key()
        manifest = b'{"files": []}'
        signature = sign_manifest(manifest, settings)
        assert signature.manifest_sha256 == hashlib.sha256(manifest).hexdigest()


class TestExcelExport:
    def _sample_records(self):
        now = dt.datetime.now(dt.timezone.utc)
        rule = RuleChangeRecord(
            rule_id="r1", rule_version=1, clause_number="4.2.b", circular_number="SEBI/HO/1/2025/1",
            is_active=True, hitl_status="NONE", compiler_version="1.0.0", created_at=now, opa_package_name="pkg",
        )
        hitl = HITLApprovalRecord(
            review_id="rev1", clause_number="4.2.b", circular_number="SEBI/HO/1/2025/1", reason_code="qualitative_directive",
            severity="blocking", status="RESOLVED", compliance_officer_id="officer.jane", review_notes=None,
            resolution_notes="Approved", flagged_at=now, resolved_at=now,
        )
        ledger = AuditTrailEntry(
            sequence_num=1, broker_id="INZ0001001", transaction_id="TXN-1", evaluated_at=now,
            circular_id="SEBI/HO/1/2025/1", section_reference="4.2.b", rule_id="r1", evaluation_result="FAIL",
            hitl_review_id=None, payload_digest="a" * 64, current_hash="b" * 64,
        )
        circular = SourceCircularRecord(
            circular_number="SEBI/HO/1/2025/1", title="Test Circular", issue_date=dt.date(2025, 7, 1),
            source_url="https://sebi.gov.in/x.pdf", raw_text_digest="c" * 64, department="MIRSD",
        )
        return [rule], [hitl], [ledger], [circular]

    def test_workbook_has_all_expected_sheets(self) -> None:
        rules, hitl, ledger, circulars = self._sample_records()
        xlsx_bytes = build_excel_workbook(period_label="Q2-2025", rule_changes=rules, hitl_approvals=hitl, ledger_entries=ledger, source_circulars=circulars)

        import openpyxl

        wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
        assert wb.sheetnames == ["Summary", "Rule Changes", "HITL Approvals", "Ledger Proof Chain", "Source Circulars"]

    def test_rule_changes_sheet_contains_the_row(self) -> None:
        rules, hitl, ledger, circulars = self._sample_records()
        xlsx_bytes = build_excel_workbook(period_label="Q2-2025", rule_changes=rules, hitl_approvals=hitl, ledger_entries=ledger, source_circulars=circulars)

        import openpyxl

        wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
        rows = list(wb["Rule Changes"].iter_rows(values_only=True))
        assert rows[0][0] == "Rule ID"
        assert rows[1][0] == "r1"
        assert rows[1][3] == "SEBI/HO/1/2025/1"

    def test_empty_data_still_produces_valid_workbook(self) -> None:
        xlsx_bytes = build_excel_workbook(period_label="Q2-2025", rule_changes=[], hitl_approvals=[], ledger_entries=[], source_circulars=[])
        import openpyxl

        wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
        assert wb["Rule Changes"].max_row == 1  # header only
